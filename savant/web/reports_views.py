
"""
    Hey, developer! Looking on the code below don't blame author.
    First version that came to me was in raw html+javascript (90Kb+161Kb)
        and I managed to rewrite it to only 70Kb (later increased due to added functionality).
    It became much more readable and flexible (trust me).
    I do apologize - It can be hard to debug/change some parts, but at the moment of creation
    it worked fast and customer was happy.
    Alexey
"""
import math
import os
import re
import io
import datetime
import json
import itertools as it
from hashlib import md5
from collections import defaultdict, OrderedDict
import dateutil.relativedelta
from tempfile import mkstemp

import openpyxl
import pdfkit
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg

from django.template.loader import render_to_string
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.core.urlresolvers import reverse
from django.db.models import Q, Sum
from django.utils import timezone
from django.conf import settings

from . import models
from . import tasks
from .utils import list2csv, formatkpi, get_fonts, replace_date_placeholders

CUR_DIR = os.path.dirname(__file__)

EMAIL_EXTRACT_REX = re.compile(r'(\S+@\S+)', re.I)

REPORT_FILTERS = [
    'site[]', 'kpi[]', 'zone[]', 'chain[]', 'hours[]', 'date', 'daterange', 'comparison',
]


def get_div(value, div):
    if not div or not value or div == 0:
        return 0
    return 1.0 * value / div


def get_reports_list(user):
    reports = []
    all_sett = user.get_all_settings_dict()
    for reportid in user.enabled_reports:
        creport = {
            'reportid': reportid,
            'urlname': "%s-report" % reportid,
            'shortname': all_sett.get(reportid, reportid),
            'name': "%s Report" % all_sett.get(reportid, reportid),
        }
        reports.append(creport)
    return reports


def get_report_url_name(reportid):
    return "%s-report" % reportid


def get_report_url(report_url_name):
    return reverse(report_url_name)


def get_known_reports():
    if not hasattr(get_known_reports, '_reports_cache'):
        get_known_reports._reports_cache = [
            PerformanceCalendarReportView,
            ExecutiveSummaryReportView,
            HourlyPerformanceReportView,
            DailyRetailTrendAnalysisView,
            PerformanceComparisonView,
            PerformanceTrendAnalysisReportView,
        ]
    return get_known_reports._reports_cache


def get_known_reports_dict():
    if not hasattr(get_known_reports_dict, '_reports_cache'):
        get_known_reports_dict._reports_cache = dict((rklass.get_reportid(), rklass) for rklass in get_known_reports())
    return get_known_reports_dict._reports_cache


def save_report_file(stream, extension, reportlog):
    if extension == 'xls':
        extension = 'xlsx'
    fpath = "%s/%s.%s" % (settings.REPORTS_STORE_DIR,
                          reportlog.report_relfilepath(),
                          extension)
    if not os.path.exists(os.path.dirname(fpath)):
        os.makedirs(os.path.dirname(fpath))
    try:
        with open(fpath, 'wb') as rff:
            rff.write(stream)
    except TypeError:
        # "a bytes-like object is required, not 'str'"
        with open(fpath, 'w') as rff:
            rff.write(stream)


class ReportsListView(TemplateView):
    template_name = 'reports/reports_list.html'

    def get_context_data(self):
        context = super().get_context_data()
        context['reports'] = get_reports_list(self.request.user)
        return context


class BaseReportView(TemplateView):
    template_name = 'reports/base_reports.html'
    data_template_name = 'reports/empty.html'
    pdf_template_name = 'reports/pdf/base_pdf.html'
    report_name = "Unknown Report"
    # export_formats = ['csv', 'pdf']
    export_formats = []
    filter_items = ['Chains', 'Sites', 'Zones', 'DateRange', 'Date', 'Hours', 'KPI']  # just all

    def kpiname(self, kpiid):
        return self.settings['kpiname%s' % kpiid.replace('%', '')]

    @classmethod
    def get_reportid(cls):
        # Some dark magic to calculate report name
        return cls.__name__.lower().replace('reportview', '').replace('view', '')

    def get_context_data(self):
        if self.request.method == 'GET':
            self.request_params = self.request.GET
        elif self.request.method == 'POST':
            self.request_params = self.request.POST
        # print(self.request_params)
        context = super().get_context_data()
        self.settings = self.request.user.get_all_settings_dict()
        self.KPI_LIST = OrderedDict(list(models.Kpi.objects.all().values_list('name', 'displayname')))
        self.KPI_NAMES_REVERSED = dict((self.kpiname(k), k) for k in self.KPI_LIST)
        self.KPI_FORMATS = dict((kpi, self.settings['kpiformat%s' % kpi.replace('%', '')]) for kpi in self.KPI_LIST)
        context['kpi_list'] = []
        for kpiname, kpidn in list(self.request.user.get_kpis().values_list('name', 'displayname')):
        # for kpiname, kpidn in list(self.request.user.enabled_kpis.all().values_list('name', 'displayname')):            
            context['kpi_list'].append((kpiname, self.kpiname(kpiname)))

        reportid = self.get_reportid()
        # context['report_name'] = self.report_name
        context['report_name'] = self.settings.get(reportid, self.report_name)
        context['report_id'] = reportid
        context['export_formats'] = self.export_formats
        context['export'] = self.request_params.get('data_format')
        context['showkpiid'] = self.request_params.get('showkpiid')
        context['report_data_tpl'] = self.data_template_name
        context['report_url'] = get_report_url_name(reportid)
        context['reportheader'] = self.settings["%s_header" % reportid]
        context['reportfooter'] = self.settings["%s_footer" % reportid]
        context['settings'] = json.dumps(self.settings)
        context['settings_raw'] = self.settings
        current_time = self.request.GET.get('current_time', '').lower()
        current_time = current_time or datetime.datetime.now().strftime('%m-%d-%Y-%H-%M')
        context['hour_min_avail'] = [0, 30]
        context['delivery_period_types'] = models.ScheduledReport.DELIVERY_PERIOD_TYPES
        context['weekdays'] = models.ScheduledReport.WEEKDAYS
        context['current_time'] = current_time
        report_date_format = self.settings.get(reportid + '_date_format', '').strip()
        if report_date_format:
            context['current_time_report'] = replace_date_placeholders(
                report_date_format,
                datetime.datetime.strptime(current_time, '%m-%d-%Y-%H-%M')
            )
        else:
            context['current_time_report'] = ''

        # ..._footer_date_format
        report_footer_date_format = self.settings.get(reportid + '_footer_date_format', '').strip()
        if report_footer_date_format:
            context['current_time_report_footer'] = replace_date_placeholders(
                report_footer_date_format,
                datetime.datetime.strptime(current_time, '%m-%d-%Y-%H-%M')
            )
        else:
            context['current_time_report_footer'] = ''

        context['initial'] = {}
        for pr in REPORT_FILTERS:
            if pr.endswith('[]'):
                context['initial'][pr.replace('[]', '')] = self.request_params.getlist(pr)
            else:
                context['initial'][pr] = self.request_params.get(pr)
        context['initial']['reporturl'] = reverse(context['report_url'])
        context['initial_json'] = json.dumps(context['initial'])
        context['open_report'] = self.request_params.get('open_report')
        if not self.is_data_request():  # full HTML
            context['defaults'] = json.dumps(models.SETTINGS_DEFAULTS)

            context['chains'] = self.request.user.get_all_chains()

            chain2site = defaultdict(list)
            sites = self.request.user.get_all_sites().values_list('id', 'name', 'chain_id', 'chain__name')
            for siteid, sitename, chainid, chainname in sites:
                chain2site[chainid].append((siteid, "%s :: %s" % (chainname, sitename)))
            context['sites'] = json.dumps(chain2site)

            site2zone = defaultdict(list)
            zones = self.request.user.get_all_zones().values_list(
                'id', 'name', 'site_id', 'site__name', 'site__chain__name')
            for zoneid, zonename, siteid, sitename, chainname in zones:
                site2zone[siteid].append((zoneid, "%s :: %s :: %s" % (chainname, sitename, zonename)))
            context['zones'] = json.dumps(site2zone)

        else:  # ajax or export
            context['report'] = self.get_report_data()
            chains = self.chains.values('name', 'id')
            sites = self.sites.values('name', 'id', 'chain__name')
            zones = self.zones.values('name', 'id', 'site__chain__name', 'site__name')
            items = {
                'Chains': ('label1', ['%s' % (e['name']) for e in chains]),
                'Sites': ('label2', ['%s :: %s' % (e['chain__name'], e['name']) for e in sites]),
                'Zones': ('label3', ['%s :: %s :: %s' %
                                     (e['site__chain__name'], e['site__name'], e['name'])
                                     for e in zones]),
                'DateRange': ('label4', [self.paramdaterange]),
                'Date': ('label4', [self.date]),
                'Hours': ('label5', ["%02d:00 - %02d:59" % (h, h) for h in self.report_hours]),
                'KPI': ('label6',
                        [self.kpiname(k) for k in context['report']['kpi'] if context['report']['kpi'][k]['enabled']]),
            }
            context['enabled_items'] = [(self.settings.get(items[i][0]), items[i][1]) for i in self.filter_items]
        return context

    def get_report_data(self):
        # to be overridden later
        retval = {
            'csvdata': [],
        }
        return retval

    def filter_reports(self, chain=True, zone=False, site=True, hours=False, daterange=True):
        # filter reports by request parameters
        params = self.request_params
        reports = self.request.user.get_all_reports()
        self.chains = models.Chain.objects.all()
        if params.getlist('chain[]'):
            l = params.getlist('chain[]')
            self.chains = self.request.user.get_all_chains(subset=l)
            reports = reports.filter(chain__in=l)
        self.sites = models.Site.objects.all()
        self.zones = models.Zone.objects.all()
        if site and params.getlist('site[]'):
            l = params.getlist('site[]')
            self.sites = self.request.user.get_all_sites(subset=l)
            reports = reports.filter(site__in=l)
        if zone and params.getlist('zone[]'):
            l = params.getlist('zone[]')
            self.zones = self.request.user.get_all_zones(subset=l)
            reports = reports.filter(zone__in=l)

        self.report_hours = []
        paramhours = params.getlist('hours[]')
        if hours and paramhours and len(paramhours) < 24:
            hours_q = Q()
            for hr in paramhours:
                hours_q |= Q(datetime__hour=hr)
                self.report_hours.append(int(hr))
            reports = reports.filter(hours_q)
        else:
            self.report_hours = list(range(24))
        # retval['hours'] = ["%02d:00 - %02d:59" % (h, h) for h in self.report_hours]

        self.date = params.get('date')
        self.paramdaterange = params.get('daterange')
        if daterange and self.paramdaterange:
            dstart, dend = self.paramdaterange.split(' - ')
            dstart = timezone.make_aware(datetime.datetime.strptime(dstart, "%m/%d/%Y"))
            dend = timezone.make_aware(datetime.datetime.strptime(dend, "%m/%d/%Y"))
            reports = reports.filter(datetime__gte=dstart)
            reports = reports.filter(datetime__lte=dend)

        return reports

    def aggregate_reports(self, reports, values, order_field=None):
        result = []

        ann_fields = {}
        ann_fields['kpi_footfall'] = Sum('visitors_in')
        ann_fields['kpi_sales'] = Sum('sales')
        ann_fields['kpi_trans'] = Sum('transactions')
        ann_fields['kpi_units'] = Sum('items')
        ann_fields['kpi_staff'] = Sum('associates')

        reports = reports.values(*values)
        if ann_fields:
            reports = reports.annotate(**ann_fields)
        if order_field:
            reports = reports.order_by(order_field)

        for row in list(reports):
            row['kpi_conv'] = get_div(row['kpi_trans'], row['kpi_footfall']) * 100
            row['kpi_atv'] = get_div(row['kpi_sales'], row['kpi_trans'])
            row['kpi_upt'] = get_div(row['kpi_units'], row['kpi_trans'])
            row['kpi_acv'] = get_div(row['kpi_sales'], row['kpi_footfall'])
            row['kpi_ctsr'] = get_div(row['kpi_footfall'], row['kpi_staff'])
            row['kpi_upc'] = get_div(row['kpi_units'], row['kpi_footfall'])
            row['kpi_sps'] = get_div(row['kpi_sales'], row['kpi_staff'])
            result.append(row)

        # print("Selected %s reports" % (len(result)))
        return result

    def get_kpis_initial(self, only=None):
        kpis = self.request_params.getlist('kpi[]') or ['FOOTFALL']
        # if 'FOOTFALL' not in kpis:
        #     kpis.append('FOOTFALL')
        if only:
            kpis = only
        kpi_cnt = 0
        kpi_sum = OrderedDict((k, {
            'sum': 0,
            'css': k.replace('%', '').lower(),
            'format': self.KPI_FORMATS[k],
            'id': "kpi_%s" % k.replace('%', '').lower(),
            'name': k,
            'fullname': self.kpiname(k),
            'description': self.KPI_LIST[k],
            'enabled': k in kpis,
            'week': {},
            'weekday': {},
        }) for k in self.KPI_LIST)
        return kpi_sum

    def calc_kpis(self, kpis, key='sum', reduce=False):
        if isinstance(kpis['FOOTFALL'][key], dict):
            for k in kpis['FOOTFALL'][key].keys():
                kpis['%CONV'][key][k] = get_div(kpis['TRANS'][key][k], kpis['FOOTFALL'][key][k]) * 100
                kpis['ATV'][key][k] = get_div(kpis['SALES'][key][k], kpis['TRANS'][key][k])
                kpis['UPT'][key][k] = get_div(kpis['UNITS'][key][k], kpis['TRANS'][key][k])
                kpis['ACV'][key][k] = get_div(kpis['SALES'][key][k], kpis['FOOTFALL'][key][k])
                kpis['CTSR'][key][k] = get_div(kpis['FOOTFALL'][key][k], kpis['STAFF'][key][k])
                kpis['UPC'][key][k] = get_div(kpis['UNITS'][key][k], kpis['FOOTFALL'][key][k])
                kpis['SPS'][key][k] = get_div(kpis['SALES'][key][k], kpis['STAFF'][key][k])
        else:
            kpis['%CONV'][key] = get_div(kpis['TRANS'][key], kpis['FOOTFALL'][key]) * 100
            kpis['ATV'][key] = get_div(kpis['SALES'][key], kpis['TRANS'][key])
            kpis['UPT'][key] = get_div(kpis['UNITS'][key], kpis['TRANS'][key])
            kpis['ACV'][key] = get_div(kpis['SALES'][key], kpis['FOOTFALL'][key])
            kpis['CTSR'][key] = get_div(kpis['FOOTFALL'][key], kpis['STAFF'][key])
            kpis['UPC'][key] = get_div(kpis['UNITS'][key], kpis['FOOTFALL'][key])
            kpis['SPS'][key] = get_div(kpis['SALES'][key], kpis['STAFF'][key])
        if reduce:
            kpis = OrderedDict((i, kpis[i]) for i in kpis if kpis[i]['enabled'])
        return kpis

    def get_color(self, kpiname, use_reversed=False):
        if not hasattr(self, '_usettings'):
            self._usettings = self.request.user.get_all_settings_dict()
        kname = kpiname
        if use_reversed:
            kname = self.KPI_NAMES_REVERSED.get(kpiname)
        color = self._usettings.get(kname)
        return color

    def is_data_request(self):
        if self.request.is_ajax():
            return True
        if self.request.method == 'GET' and 'data_format' in self.request.GET or \
                self.request.method == 'POST' and 'data_format' in self.request_params:
            return True
        return False

    def get_xls_styles(self):
        styles = {}
        styles['header_fill'] = openpyxl.styles.PatternFill(start_color='006a72', end_color='006a72', fill_type='solid')
        styles['header_font'] = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE, bold=True)
        styles['cell_font'] = openpyxl.styles.Font(color=openpyxl.styles.colors.BLACK)
        return styles

    def write_xls(self, book, context):
        pass

    def xls_draw_chart(self, book, chartdata, chart_name=None, sheet_name=None):
        if not chartdata:
            return
        sheet_name = sheet_name or 'Charts'
        book.create_sheet(sheet_name, len(book.worksheets))
        sheet = book.get_sheet_by_name(sheet_name)
        c1 = openpyxl.chart.LineChart()
        c1.title = chart_name
        # c1.style = 30
        startrow = 200
        # startrow = 1
        sheet._current_row = startrow - 1  # small dirty hack
        whitefont = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE)
        sheet.append(chartdata['categories'])
        for c in range(len(chartdata['categories'])):
            sheet.cell(column=c + 1, row=startrow).font = whitefont
        refcategory = openpyxl.chart.Reference(sheet, min_col=1, min_row=startrow,
                                               max_col=len(chartdata['categories']), max_row=startrow)
        for rownum, row in enumerate(chartdata['series']):
            current_row = startrow + rownum + 1
            # sheet.append(row['data'])
            for colnum, coldata in enumerate(row['data']):
                cell = sheet.cell(column=colnum + 1, row=current_row)
                # if isinstance(coldata, str):
                #     if coldata.endswith('%'):
                #         coldata = coldata[:-1]
                #     if '.' in coldata:
                #         coldata = float(coldata)
                #     else:
                #         coldata = int(coldata)
                cell.value = coldata
                cell.font = whitefont
            refdata = openpyxl.chart.Reference(sheet, min_col=1, min_row=current_row,
                                               max_col=len(row['data']), max_row=current_row)
            series = openpyxl.chart.Series(refdata, title=row['name'])
            scolor = row['color']
            if scolor.startswith('#'):
                scolor = scolor[1:]
            else:
                scolor = None
            if scolor:
                series.graphicalProperties.line.solidFill = scolor.upper()
            series.smooth = True
            c1.append(series)
        c1.set_categories(refcategory)
        c1.width = 35
        c1.height = 15
        sheet.add_chart(c1, "A1")

    MARKERS = ['o', 'v', 's', '*', 'x']

    def draw_chart_png(self, data):
        figure = Figure()
        canvas = FigureCanvasAgg(figure)
        figure.suptitle(data.get('name', 'Chart'), fontsize='xx-large')
        ax = figure.add_subplot(1, 1, 1)
        ax.set_xticks(range(len(data['categories'])))
        ax.axes.set_xticklabels(data['categories'])
        ax.spines['right'].set_visible(False)
        ax.spines['top'].set_visible(False)
        ax.yaxis.set_ticks_position('left')
        ax.xaxis.set_ticks_position('bottom')
        for series_data, marker in zip(data['series'],
                                       it.cycle(self.MARKERS)):
            ax.plot(
                series_data['data'][:len(data['categories'])],
                color=series_data['color'],
                linewidth=2.5,
                label=series_data['name'],
                marker=marker,
                markerfacecolor=series_data['color'],
                markeredgecolor=series_data['color'],
                markeredgewidth=2
            )
        max_label_legnth = max(
            len(series_data['name'])
            for series_data in data['series']
        )
        ax.legend(
            loc='upper center',
            bbox_to_anchor=(0.5, -0.05),
            frameon=False,
            fontsize='xx-small',
            ncol=int(math.ceil(8. / max_label_legnth * 11.)),
            markerscale=0.8
        )

        # Rotate xlabels in case if they are not numbers
        # and adjust legend position
        xlabels = ax.get_xaxis().get_ticklabels()
        try:
            float(xlabels[0].get_text())
        except ValueError:
            for xlabel in xlabels:
                xlabel.set_rotation(-45)
                xlabel.set_size('x-small')
            max_xlabel_size = max(
                len(xlabel.get_text()) for xlabel in xlabels
            )
            ax.get_legend().set_bbox_to_anchor((
                0.5, - 0.05 - (0.01 * max_xlabel_size)
            ))

        file = io.BytesIO()
        canvas.print_figure(
            file,
            format='png',
            dpi=120,
            bbox_inches='tight'
        )

        return file.getvalue()

    def create_pdf(self, context, chartfiles=None):
        context['fonts'] = get_fonts(self.settings)
        sRoot = settings.STATICFILES_DIRS[0]
        context['static_root'] = sRoot
        if self.request.user.logo.name:
            context['logo_file'] = self.request.user.logo.path
        else:
            context['logo_file'] = os.path.join(sRoot, 'web/img/logo_default.png')
        data = context['report'].get('chartdata_raw')
        if data:
            if isinstance(data, dict):
                data = [data]
            context['chartfiles'] = chartfiles = []
            for figure_data in data:
                if 'report_name' in context:
                    figure_data.setdefault('name', context['report_name'])
                image = self.draw_chart_png(figure_data)
                fd, path = mkstemp(prefix='tempchart_', suffix='.png')
                file = os.fdopen(fd, mode='w+b')
                file.write(image)
                chartfiles.append(path)

        html = render_to_string(self.pdf_template_name, context)
        # with open('/tmp/html2pdf.html', 'w') as dbghtml:
        #     dbghtml.write(html)  # debug
        cssfiles = [
            os.path.join(sRoot, 'web/css/bootstrap.min.css'),
            os.path.join(sRoot, 'web/css/styles.css'),
            os.path.join(sRoot, 'web/css/reports.css'),
            os.path.join(sRoot, 'web/css/print.css'),
        ]
        # jsfiles = [
        #     os.path.join(sRoot, 'web/js/reports_styles.js'),
        # ]
        options = {
            'page-size': 'Letter',
            'no-outline': None,
            'debug-javascript': None,
            'enable-javascript': None,
            # 'run-script': jsfiles,
            # 'orientation': 'Portrait',
            'quiet': ''
        }

        if context['reportfooter']:
            options['footer-center'] = context['reportfooter']
            
        if context['current_time_report_footer']:
            options['footer-center'] = '{} {}'.format(
                context['reportfooter'],
                context['current_time_report_footer'])

        config = None
        config = pdfkit.configuration(wkhtmltopdf=settings.WKHTMLTOPDF_BIN)
        pdffile = pdfkit.from_string(html, False, options=options, css=cssfiles, configuration=config)
        return pdffile, chartfiles

    def create_xls(self, context):
        book = openpyxl.Workbook()
        sheet = book.active
        # create summary sheet
        book.create_sheet("Summary", 0)
        summary_sheet = book.get_sheet_by_name("Summary")
        rprtname = context['report_name']
        cell1 = summary_sheet.cell(row=1, column=1)
        cell1.value = "%s Report" % rprtname
        cell1.font = openpyxl.styles.Font(color=openpyxl.styles.colors.BLACK, bold=True)
        if context.get('current_time_report'):
            summary_sheet.cell(row=2, column=1).value = "Date"
            summary_sheet.cell(row=2, column=2).value = context['current_time_report']
        summary_sheet.cell(row=3, column=1).value = "Filter values:"
        crow = 3
        for item in context['enabled_items']:
            crow += 1
            summary_sheet.cell(row=crow, column=1).value = item[0]
            fval = item[1]
            if isinstance(fval, list):
                fval = ', '.join(fval)
            summary_sheet.cell(row=crow, column=2).value = fval
        # cell.font = styles['header_font']
        # cell.fill = styles['header_fill']
        for dimension in summary_sheet.column_dimensions.values():
            dimension.auto_size = True

        book.active = 1
        if len(context['report_name']) > 31:
            # try to cut last chars. Only 31 chars allowed in the sheet name
            if rprtname.lower().endswith(' report'):
                rprtname = rprtname[:-7]
        sheet.title = rprtname[:31]
        self.write_xls(book, context)
        book.active = 0
        chartdata = context['report'].get('chartdata_raw')
        if chartdata:
            if isinstance(chartdata, dict):
                chart_name = context['report_name']
                self.xls_draw_chart(book, chartdata, chart_name, sheet_name=None)
            else:
                for chart in chartdata:
                    chart_name = chart['name']
                    self.xls_draw_chart(book, chart, chart_name, sheet_name="Chart - " + chart_name)
        buffer = openpyxl.writer.excel.save_virtual_workbook(book)
        return buffer

    def get(self, request):
        if not self.is_data_request():
            return super().get(request)
        content_disposition = "inline"
        # content_disposition = "attachment"
        context = self.get_context_data()
        data_format = self.request.GET.get('data_format', '').lower()
        file_name = "%s-%s" % (context['report_id'], context['current_time'])
        filedata = ""
        reportlog = models.ReportLog(user=self.request.user,
                                     report_id=context['report_id'],
                                     name=context['report_name'],
                                     is_manual=True,
                                     is_email=False,
                                     parameters='',
                                     calendar_date_selected=self.paramdaterange or self.date or '',
                                     )
        if data_format == 'print':  # or (data_format == 'pdf' and 'pdf' in self.export_formats):
            context['print'] = True
            self.template_name = 'reports/base_print_report.html'
            return self.render_to_response(context)
        elif data_format == 'csv' and 'csv' in self.export_formats:
            filedata = list2csv(context['report']['csvdata'])
            file_name += '.csv'
            response = HttpResponse(filedata, content_type='text/csv')
            response['Content-Disposition'] = '%s; filename="%s"' % (content_disposition, file_name)
            reportlog.is_csv = True
            reportlog.save()
            save_report_file(filedata, 'csv', reportlog)
            return response
        elif data_format == 'xls' and 'xls' in self.export_formats:
            filedata = self.create_xls(context)
            file_name += '.xlsx'
            response = HttpResponse(filedata, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = '%s; filename="%s"' % (content_disposition, file_name)
            reportlog.is_xls = True
            reportlog.save()
            save_report_file(filedata, 'xls', reportlog)
            return response
        elif data_format == 'pdf' and 'pdf' in self.export_formats:
            reporturl = self.request.build_absolute_uri()
            urlhash = md5(reporturl.encode()).hexdigest()
            chartfiles = self.request.session.get(urlhash) or []
            filedata, tempfiles = self.create_pdf(context)
            if content_disposition == "attachment":
                # We CANNOT delete the files if 'content_disposition' is "inline":
                # the report get opened in the browser and after refresh we lose the charts
                if tempfiles:
                    tasks.delete_files.apply_async(countdown=172800, kwargs={'files_list': tempfiles})
                if chartfiles:
                    # delete in 2 days (60 * 60 * 24 * 2)
                    tasks.delete_files.apply_async(countdown=172800, kwargs={'files_list': chartfiles})
                    # for cfile in chartfiles:
                    #     try:
                    #         os.unlink(cfile)
                    #     except:
                    #         pass
                    # del self.request.session[urlhash]
            file_name += '.pdf'
            response = HttpResponse(filedata, content_type='application/pdf')
            response['Content-Disposition'] = '%s; filename="%s"' % (content_disposition, file_name)
            reportlog.is_pdf = True
            reportlog.save()
            save_report_file(filedata, 'pdf', reportlog)
            return response
        elif data_format == 'html' and 'html' in self.export_formats:
            self.template_name = self.data_template_name
            return self.render_to_response(context)
        return HttpResponse(status=400, content="Invalid parameters set")

    def post(self, request):
        if not self.is_data_request():
            return super().post(request)
        # print(self.request)
        # print(self.request.POST)
        context = self.get_context_data()
        if 'savechart' in self.request.POST and 'chartdata[]' in self.request.POST:
            # Get rendered SVG charts from browser and save them on disk to insert to PDF report later
            # Some 'dirty' solution. We need charts in PDF, but will not render them on the backend.
            # But as we already have them on frontend - we can post them when user requests a report export.
            charts = self.request.POST.getlist('chartdata[]')
            reporturl = self.request.POST.get('reporturl')
            urlhash = md5(reporturl.encode()).hexdigest()
            self.request.session[urlhash] = []
            for chart in charts:
                fhandle, filename = mkstemp(prefix='tempchart_', suffix='.svg')
                os.write(fhandle, chart.encode())
                os.close(fhandle)
                self.request.session[urlhash].append(filename)
            return HttpResponse("OK")
        elif 'emailreport' in self.request.POST:
            # print(self.request.POST)
            emailreport = self.request.POST.get('emailreport')
            formats = self.request.POST.getlist('formats[]')
            reportfiles = []
            file_name = "%s-%s" % (context['report_id'], context['current_time'])
            emaildateformat = self.request.POST.get('emaildateformat') or '%Y-%m-%d'
            # emaildateformat = 'Created at {}'.format(emaildateformat)
            subject = self.request.POST.get('emailsubject') or 'Report'
            subject = '{}. {}'.format(subject,emaildateformat)            
            subject = replace_date_placeholders(subject, datetime.datetime.now())
            body = self.request.POST.get('emailbody') or 'Report'
            body = '{}. {}'.format(body,emaildateformat)  
            body = replace_date_placeholders(body, datetime.datetime.now())
            emailer = models.Emailer(type=models.Emailer.TEXT,
                                     subject=subject,
                                     body=body,
                                     email_from=settings.DEFAULT_FROM_EMAIL,
                                     )
            emailer.email_to_many = ','.join(EMAIL_EXTRACT_REX.findall(self.request.POST.get('emailto')))
            emailer.email_cc_many = ','.join(EMAIL_EXTRACT_REX.findall(self.request.POST.get('emailcc')))
            emailer.email_bcc_many = ','.join(EMAIL_EXTRACT_REX.findall(self.request.POST.get('emailbcc')))
            if not (emailer.email_to_many or emailer.email_cc_many or emailer.email_bcc_many):
                return HttpResponse("Error: Recipients should be set!")
            exclude_fields = ('emailto', 'emailcc', 'emailbcc', 'emailsubject', 'emailbody', 'formats[]', 'emailreport')
            rparams = context['initial_json']
            reportlog = models.ReportLog(user=self.request.user,
                                         report_id=context['report_id'],
                                         name=context['report_name'],
                                         is_manual=emailreport == 'manual',
                                         is_email=True,
                                         parameters=rparams,
                                         calendar_date_selected=self.paramdaterange or self.date or '',
                                         )
            if 'csv' in formats:
                filedata = list2csv(context['report']['csvdata'])
                fhandle, datafile = mkstemp(prefix='tempemailreport_', suffix='.csv')
                os.write(fhandle, filedata.encode())
                os.close(fhandle)
                reportfiles.append({'format': 'csv', 'filename': file_name + '.csv', 'datafile': datafile})
                reportlog.is_csv = True
            if 'xls' in formats:
                filedata = self.create_xls(context)
                fhandle, datafile = mkstemp(prefix='tempemailreport_', suffix='.xlsx')
                os.write(fhandle, filedata)
                os.close(fhandle)
                reportfiles.append({'format': 'xls', 'filename': file_name + '.xlsx', 'datafile': datafile})
                reportlog.is_xls = True
            if 'pdf' in formats:
                reporturl = self.request.POST.get('reporturl')
                urlhash = md5(reporturl.encode()).hexdigest()
                chartfiles = self.request.session.get(urlhash) or []
                filedata, tempfiles = self.create_pdf(context)
                fhandle, datafile = mkstemp(prefix='tempemailreport_', suffix='.pdf')
                os.write(fhandle, filedata)
                os.close(fhandle)
                if tempfiles:
                    tasks.delete_files.apply_async(kwargs={'files_list': tempfiles})
                if chartfiles:
                    tasks.delete_files.apply_async(kwargs={'files_list': chartfiles})
                    del self.request.session[urlhash]
                    # for cfile in chartfiles:
                    #     try:
                    #         os.unlink(cfile)
                    #     except:
                    #         pass
                    # del self.request.session[urlhash]
                reportfiles.append({'format': 'pdf', 'filename': file_name + '.pdf', 'datafile': datafile})
                reportlog.is_pdf = True
            if not reportfiles:
                return HttpResponse("Error! Please select an attachment!")
            emailer.set_attachment_pathes(reportfiles)
            emailer.save()
            reportlog.email = emailer
            reportlog.save()
            # WTF do I do here? Simply put the file to another place
            for ff in reportfiles:
                with open(ff['datafile'], 'rb') as rfr:
                    filedata = rfr.read()
                    save_report_file(filedata, ff['format'], reportlog)
            return HttpResponse("OK")
        self.template_name = self.data_template_name
        return self.render_to_response(context)


class PerformanceComparisonView(BaseReportView):
    template_name = 'reports/performancecomparison.html'
    data_template_name = 'reports/performancecomparison_data.html'
    pdf_template_name = 'reports/pdf/performancecomparison.html'
    report_name = "Performance Comparison Report"
    export_formats = ['pdf', 'xls', 'csv']
    filter_items = ['Chains', 'Sites', 'KPI', 'Zones', 'DateRange', 'Hours']

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # make copy: below we modify it (to deal with instance-wide var, not class-wide)
        self.filter_items = self.filter_items.copy()

    def write_xls(self, book, context):
        csvdata = context['report']['csvdata']
        styles = self.get_xls_styles()
        sheet = book.active
        alcenter = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        # writing first row
        for si, dmn in enumerate(context['report']['iterdomains']):
            sname = "%s-%s" % (dmn['chain__name'], dmn['name'])
            sheet.merge_cells(start_row=1, start_column=si * context['report']['kpi_len'] + 2,
                              end_row=1, end_column=(si + 1) * context['report']['kpi_len'] + 1)
            cell = sheet.cell(row=1, column=si * context['report']['kpi_len'] + 2)
            cell.value = sname
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = alcenter
        cur_row = 2
        default_color = ""
        for cri, csvrow in enumerate(csvdata):
            hour = csvrow[0]
            for cci, csvcell in enumerate(csvrow):
                if cci == 0:
                    # TimePeriod column
                    cell = sheet.cell(row=cur_row + cri, column=cci + 1)
                    cell.value = csvcell
                    cell.font = styles['header_font']
                    cell.fill = styles['header_fill']
                    continue
                hdrnames = csvdata[0][cci].split(' :: ')
                kpiname = hdrnames[-1]
                # book.active = sheet
                cell = sheet.cell(row=cur_row + cri, column=cci + 1)
                cell.font = styles['cell_font']
                cell.fill = styles['header_fill']
                if cri == 0:
                    cell.font = styles['header_font']
                    cell.fill = styles['header_fill']
                    cell.value = kpiname
                else:
                    color = self.get_color(kpiname, use_reversed=True)
                    if color.startswith('#'):
                        color = color[1:]
                    cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.value = csvcell

    def aggregate_reports(self, reports, is_site=True):
        rep_values = ['chain_id', 'site_id', 'site_name', 'chain_name', 'hour']
        if not is_site:
            rep_values.extend(['zone_name', 'zone_id'])
        reports = reports\
            .extra(select={'hour': "date_part('hour', datetime)"})
        return super().aggregate_reports(reports, values=rep_values, order_field='hour')

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()
        is_site = params.get('comparison', '').lower() == 'true'
        retval['is_site'] = is_site

        reports = self.filter_reports(chain=True, zone=not is_site, site=True, hours=True, daterange=True)
        if is_site:
            kpi_sum = self.get_kpis_initial()
            self.filter_items.remove('Zones')
        else:
            kpi_sum = self.get_kpis_initial(only=['FOOTFALL'])
            self.filter_items.remove('KPI')

        agg_reports = self.aggregate_reports(reports, is_site=is_site)

        chartdata = {'categories': ["%02d:00 - %02d:59" % (h, h) for h in self.report_hours], 'series': OrderedDict()}
        data = dict((h, defaultdict(dict)) for h in self.report_hours)
        for kname, kpi in list(kpi_sum.items()):
            for row in agg_reports:
                value = row.get(kpi['id']) or 0
                if row['hour'] not in data:
                    data[row['hour']] = {}
                if is_site:
                    if row['site_id'] not in data[row['hour']]:
                        data[row['hour']][row['site_id']] = {}
                    data[row['hour']][row['site_id']][kpi['id']] = value
                else:
                    rk = '%s-%s' % (row['site_id'], row['zone_id'])
                    if rk not in data[row['hour']]:
                        data[row['hour']][rk] = {}
                    data[row['hour']][rk][kpi['id']] = value
                kpi_sum[kname]['sum'] += value

        retval['hours'] = self.report_hours
        retval['data'] = data
        if is_site:
            retval['iterdomains'] = self.sites.order_by('chain__name', 'name').values('name', 'id', 'chain__name')
        else:
            retval['iterdomains'] = []
            izones = self.zones.order_by('site__chain__name', 'site__name', 'name').\
                values('name', 'id', 'site_id', 'site__chain__name', 'site__name')
            for z in izones:
                retval['iterdomains'].append({'name': "%s :: %s" % (z['site__name'], z['name']),
                                              'id': "%s-%s" % (z['site_id'], z['id']),
                                              'chain__name': z['site__chain__name']})
        retval['sites'] = self.sites.order_by('chain__name', 'name').values('name', 'id', 'chain__name')
        retval['zones'] = self.zones.order_by('site__chain__name', 'site__name', 'name').\
            values('name', 'id', 'site__chain__name', 'site__name')

        retval['kpi'] = self.calc_kpis(kpi_sum, reduce=False)
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])

        # prepare CSV and chart data
        csvheader = ['TimePeriod']
        for dmn in retval['iterdomains']:
            for kname, kpi in list(kpi_sum.items()):
                if kpi['enabled']:
                    csvheader.append("%s :: %s :: %s" % (dmn['chain__name'], dmn['name'], kpi['fullname']))
        retval['csvdata'].append(csvheader)
        for hour in data.keys():
            csvrow = ["%02d:00 - %02d:59" % (hour, hour)]
            for dmn in retval['iterdomains']:
                for kname, kpi in list(kpi_sum.items()):
                    if kpi['enabled']:
                        value = data[hour].get(dmn['id'], {}).get(kpi['id']) or 0
                        csvrow.append(formatkpi(value, kpi['format']))
                        srname = "%s :: %s" % (dmn['chain__name'], dmn['name'])
                        if is_site:
                            srname = "%s :: %s :: %s" % (dmn['chain__name'], dmn['name'], kpi['fullname'])
                        else:
                            srname = "%s :: %s" % (dmn['chain__name'], dmn['name'])
                        if srname not in chartdata['series']:
                            chartdata['series'][srname] = {'name': srname, 'type': 'spline', 'data': [],
                                                           'color': self.get_color(kname)}
                        chartdata['series'][srname]['data'].append(value)
            retval['csvdata'].append(csvrow)
        chartdata['series'] = list(chartdata['series'].values())  # from dict to list
        retval['chartdata_raw'] = chartdata
        retval['chartdata'] = json.dumps(chartdata)

        return retval


class DailyRetailTrendAnalysisView(BaseReportView):
    template_name = 'reports/dailyretailtrendanalysis.html'
    data_template_name = 'reports/dailyretailtrendanalysis_data.html'
    pdf_template_name = 'reports/pdf/dailyretailtrendanalysis.html'
    report_name = "Daily Retail Trend Analysis Report"
    export_formats = ['pdf', 'xls', 'csv']
    filter_items = ['Chains', 'Sites', 'Date', 'KPI']

    def write_xls(self, book, context):
        csvdata = context['report']['csvdata']
        sheet = book.active
        cur_row = 1
        default_color = ""
        styles = self.get_xls_styles()
        for ikpi in range(int(len(csvdata) / 4)):
            kpiname = csvdata[ikpi * 4][0]
            color = self.get_color(kpiname, use_reversed=True)
            if color.startswith('#'):
                color = color[1:]
            font = styles['cell_font']
            fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            for subrow in range(4):
                for col in range(5):
                    cell = sheet.cell(row=ikpi * 4 + subrow + 1, column=col + 1)
                    cell.value = csvdata[ikpi * 4 + subrow][col]
                    cell.font = font
                    cell.fill = fill

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()

        reports = self.filter_reports(chain=True, zone=False, site=True, hours=False, daterange=False)
        kpi_sum = self.get_kpis_initial()

        current_day = timezone.make_aware(datetime.datetime.strptime(params.get('date'), "%m/%d/%Y"))
        past_day = current_day - datetime.timedelta(days=1)
        past_week = current_day - datetime.timedelta(days=7)
        past_month = current_day - dateutil.relativedelta.relativedelta(months=1)
        past_year = current_day - dateutil.relativedelta.relativedelta(years=1)

        for keyname, rdate in [
                ('current_day', current_day),
                ('past_day', past_day),
                ('past_week', past_week),
                ('past_month', past_month),
                ('past_year', past_year),
        ]:
            drep = reports.filter(
                datetime__range=(timezone.make_aware(datetime.datetime.combine(rdate, datetime.time.min)),
                                 timezone.make_aware(datetime.datetime.combine(rdate, datetime.time.max))))
            drep = drep\
                .extra(select={'wday': "datetime::date"})
            agg_reports = self.aggregate_reports(drep, values=['wday'])
            for kname, kpi in list(kpi_sum.items()):
                kpi_sum[kname][keyname] = 0
                kpi_sum[kname]['%s_perc' % keyname] = "%.02f" % 0
                for row in agg_reports:
                    value = row.get(kpi['id']) or 0
                    kpi_sum[kname][keyname] += value
                    if keyname != 'current_day':
                        variance = 0
                        if value != 0:
                            variance = (kpi_sum[kname]['current_day'] - value) / value
                            variance *= 100
                        kpi_sum[kname]['%s_perc' % keyname] = "%.02f" % variance

        for key in ('current_day', 'past_day', 'past_week', 'past_month', 'past_year'):
            kpi_sum = self.calc_kpis(kpi_sum, key=key, reduce=False)
        kpi_sum = self.calc_kpis(kpi_sum, reduce=False)

        for kname, kpi in list(kpi_sum.items()):
            if not kpi['enabled']:
                continue
            retval['csvdata'].append([kpi['fullname'], kpi['fullname'],
                                     kpi['fullname'], kpi['fullname'], kpi['fullname']])
            retval['csvdata'].append(['Current Day', 'Past Day', 'Past Week', 'Past Month', 'Past Year'])
            retval['csvdata'].append([
                formatkpi(float(kpi['current_day']), kpi['format']),
                formatkpi(float(kpi['past_day']), kpi['format']),
                formatkpi(float(kpi['past_week']), kpi['format']),
                formatkpi(float(kpi['past_month']), kpi['format']),
                formatkpi(float(kpi['past_year']), kpi['format']),
            ])
            retval['csvdata'].append([
                '',
                formatkpi(float(kpi['past_day_perc']), kpi['format']),
                formatkpi(float(kpi['past_week_perc']), kpi['format']),
                formatkpi(float(kpi['past_month_perc']), kpi['format']),
                formatkpi(float(kpi['past_year_perc']), kpi['format']),
            ])

        retval['kpi'] = kpi_sum
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])

        return retval


class ExecutiveSummaryReportView(BaseReportView):
    template_name = 'reports/executivesummary.html'
    data_template_name = 'reports/executivesummary_data.html'
    pdf_template_name = 'reports/pdf/executivesummary.html'
    report_name = "Executive Summary Report"
    export_formats = ['pdf', 'xls', 'csv']
    filter_items = ['Chains', 'Sites', 'DateRange', 'KPI']

    def write_xls(self, book, context):
        csvdata = context['report']['csvdata']
        sheet = book.active
        cur_row = 1
        default_color = ""
        styles = self.get_xls_styles()
        for cri, csvrow in enumerate(csvdata):
            for cci, csvcell in enumerate(csvrow):
                cell = sheet.cell(row=cur_row + cri, column=cci + 1)
                cell.font = styles['cell_font']
                cell.fill = styles['header_fill']
                if cri == 0:
                    cell.font = styles['header_font']
                else:
                    kpiname = csvdata[0][cci]
                    color = self.get_color(kpiname, use_reversed=True)
                    if color.startswith('#'):
                        color = color[1:]
                    cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.value = csvcell

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()
        kpi_list = None
        is_dahsboard = 'dashboard' in self.request.GET
        if is_dahsboard:
            kpi_list = self.request.user.get_kpis().values_list('name', flat=True)
            # kpi_list = self.request.user.enabled_kpis.all().values_list('name', flat=True)            
            self.request_params = self.request_params.copy()
            today = datetime.date.today().strftime('%m/%d/%Y')
            self.request_params['daterange'] = '%s - %s' % (today, today)

        reports = self.filter_reports(chain=True, zone=False, site=True, hours=False, daterange=True)
        kpi_sum = self.get_kpis_initial(only=kpi_list)

        agg_reports = self.aggregate_reports(reports, values=['day'])
        for kname, kpi in list(kpi_sum.items()):
            kpi_sum[kname]['sum'] = 0
            for row in agg_reports:
                value = row.get(kpi['id']) or 0
                kpi_sum[kname]['sum'] += value
        retval['kpi'] = self.calc_kpis(kpi_sum, reduce=False)

        retval['kpi_items'] = [(kpiname, kpi) for kpiname, kpi in retval['kpi'].items() if kpi['enabled']]
        
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])
        retval['csvdata'].append([kpi_sum[k]['fullname'] for k in kpi_sum if kpi_sum[k]['enabled']])
        retval['csvdata'].append([kpi_sum[k]['sum'] for k in kpi_sum if kpi_sum[k]['enabled']])
        retval['is_dahsboard'] = is_dahsboard

        return retval


class HourlyPerformanceReportView(BaseReportView):
    template_name = 'reports/hourlyperformance.html'
    data_template_name = 'reports/hourlyperformance_data.html'
    pdf_template_name = 'reports/pdf/hourlyperformance.html'
    report_name = "Hourly Performance Report"
    export_formats = ['pdf', 'xls', 'csv']
    filter_items = ['Chains', 'Sites', 'Zones', 'KPI', 'DateRange', 'Hours']

    def write_xls(self, book, context):
        csvdata = context['report']['csvdata']
        sheet = book.active
        cur_row = 1
        default_color = ""
        styles = self.get_xls_styles()
        for cri, csvrow in enumerate(csvdata):
            for cci, csvcell in enumerate(csvrow):
                cell = sheet.cell(row=cur_row + cri, column=cci + 1)
                cell.font = styles['cell_font']
                cell.fill = styles['header_fill']
                if cri == 0:
                    cell.font = styles['header_font']
                elif cci == 0:
                    cell.fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                else:
                    kpiname = csvdata[0][cci]
                    color = self.get_color(kpiname, use_reversed=True)
                    if color.startswith('#'):
                        color = color[1:]
                    cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.value = csvcell

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()

        reports = self.filter_reports(chain=True, zone=True, site=True, hours=True, daterange=True)
        kpi_sum = self.get_kpis_initial()

        reports = reports.extra(select={'hour': "date_part('hour', datetime)"})
        agg_reports = self.aggregate_reports(reports, values=['hour'], order_field='hour')

        chartdata = {'categories': ["%02d:00 - %02d:59" % (h, h) for h in self.report_hours], 'series': OrderedDict()}
        data = OrderedDict((h, defaultdict(dict)) for h in self.report_hours)
        retval['csvdata'].append(['TimePeriod'] + [kpi['fullname'] for kname, kpi in kpi_sum.items() if kpi['enabled']])
        for row in agg_reports:
            for kname, kpi in list(kpi_sum.items()):
                value = row.get(kpi['id']) or 0
                if row['hour'] not in data:
                    data[row['hour']] = {}
                data[row['hour']][kpi['id']] = value
                kpi_sum[kname]['sum'] += value

        for hour in data.keys():
            csvrow = ["%02d:00 - %02d:59" % (hour, hour)]
            for kname, kpi in list(kpi_sum.items()):
                if kpi['enabled']:
                    value = data[hour].get(kpi['id']) or 0
                    csvrow.append(formatkpi(value, kpi['format']))
                    srname = kname
                    if srname not in chartdata['series']:
                        chartdata['series'][srname] = {'name': self.kpiname(srname), 'type': 'spline', 'data': [],
                                                       'color': self.get_color(kname)}
                    chartdata['series'][srname]['data'].append(value)
            retval['csvdata'].append(csvrow)

        chartdata['series'] = list(chartdata['series'].values())  # from dict to list

        retval['hours'] = self.report_hours
        retval['data'] = data
        retval['chartdata_raw'] = chartdata
        retval['chartdata'] = json.dumps(chartdata)

        retval['kpi'] = self.calc_kpis(kpi_sum, reduce=False)
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])

        return retval


class PerformanceCalendarReportView(BaseReportView):
    template_name = 'reports/performancecalendar.html'
    data_template_name = 'reports/performancecalendar_data.html'
    pdf_template_name = 'reports/pdf/performancecalendar.html'
    report_name = "Performance Calendar Report"
    export_formats = ['pdf', 'xls']
    WDS = [6, 0, 1, 2, 3, 4, 5]
    filter_items = ['Chains', 'Sites', 'Zones', 'Date', 'KPI']
    """
        TODO: it seems there is a standard module 'calendar' in Python. Probably we'd better use it in this report
    """

    def write_xls(self, book, context):
        calendar = context['calendar']
        report = context['report']
        sheet = book.active
        bside = openpyxl.styles.Side(style='thick', color="000000")
        alcenter = openpyxl.styles.Alignment(horizontal='center')
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        cell = sheet.cell(row=1, column=1)
        cell.value = calendar['firstday'].strftime('%B %Y')
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = alcenter
        cur_row = calendar_row = 3
        for col, c in enumerate(('Week#', 'Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Week Total')):
            cell = sheet.cell(row=cur_row, column=col + 1)
            cell.value = c
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = alcenter
        for wn, week in enumerate(calendar['weeks']):
            cur_row += 1
            cell = sheet.cell(row=cur_row, column=1)
            cell.value = "%s" % (wn + calendar['firstweek'])
            for col, day in enumerate(week):
                if not day['curmonth']:
                    continue
                cell = sheet.cell(row=cur_row, column=col + 2)
                cell.value = day['day']
                cell.alignment = alcenter
                cell.border = openpyxl.styles.Border(top=bside, left=bside, right=bside)
                cell.font = openpyxl.styles.Font(bold=True)
                # writing daily kpi values
                krow = -1
                for kpi in report['kpi'].values():
                    if not kpi['enabled']:
                        continue
                    krow += 1
                    cell = sheet.cell(row=cur_row + krow + 1, column=col + 2)
                    cell.value = formatkpi(report['data'].get(day['day'], {}).get(kpi['id'], 0), kpi['format'])
                    color = self.get_color(kpi['name'])
                    if color.startswith('#'):
                        color = color[1:]
                    cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.border = openpyxl.styles.Border(left=bside, right=bside)
                cell = sheet.cell(row=cur_row + krow + 1, column=col + 2)
                cell.border = openpyxl.styles.Border(bottom=bside, left=bside, right=bside)

            # writing weekly total kpi values
            cell = sheet.cell(row=cur_row, column=col + 3)
            cell.border = openpyxl.styles.Border(top=bside, left=bside, right=bside)
            krow = -1
            for kpi in report['kpi'].values():
                if not kpi['enabled']:
                    continue
                krow += 1
                cell = sheet.cell(row=cur_row + krow + 1, column=col + 3)
                cell.value = formatkpi(kpi['week'].get(wn + calendar['firstweek'], 0), kpi['format'])
                color = self.get_color(kpi['name'])
                if color.startswith('#'):
                    color = color[1:]
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.border = openpyxl.styles.Border(left=bside, right=bside)
            cell = sheet.cell(row=cur_row + krow + 1, column=col + 3)
            cell.border = openpyxl.styles.Border(bottom=bside, left=bside, right=bside)

            cur_row += report['kpi_len']

        # Day total values
        cell = sheet.cell(row=cur_row + 1, column=1)
        cell.value = "Day Total"
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = alcenter
        for col, wd in enumerate(calendar['weekdays']):
            krow = -1
            for kpi in report['kpi'].values():
                if not kpi['enabled']:
                    continue
                krow += 1
                cell = sheet.cell(row=cur_row + krow + 1, column=col + 2)
                cell.value = formatkpi(kpi['weekday'].get(wd, 0), kpi['format'])
                color = self.get_color(kpi['name'])
                if color.startswith('#'):
                    color = color[1:]
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                if krow == 0:
                    cell.border = openpyxl.styles.Border(top=bside, left=bside, right=bside)
                else:
                    cell.border = openpyxl.styles.Border(left=bside, right=bside)
            cell = sheet.cell(row=cur_row + krow + 1, column=col + 2)
            cell.border = openpyxl.styles.Border(bottom=bside, left=bside, right=bside)

        # total values
        krow = -1
        for kpi in report['kpi'].values():
            if not kpi['enabled']:
                continue
            krow += 1
            cell = sheet.cell(row=cur_row + krow + 1, column=col + 3)
            cell.value = formatkpi(kpi['sum'], kpi['format'])
            color = self.get_color(kpi['name'])
            if color.startswith('#'):
                color = color[1:]
            cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            if krow == 0:
                cell.border = openpyxl.styles.Border(top=bside, left=bside, right=bside)
            else:
                cell.border = openpyxl.styles.Border(left=bside, right=bside)
        cell = sheet.cell(row=cur_row + krow + 1, column=col + 3)
        cell.border = openpyxl.styles.Border(bottom=bside, left=bside, right=bside)

        # draw legend
        cur_row = calendar_row
        for wn in range(len(calendar['weeks']) + 1):
            if wn < len(calendar['weeks']):
                cur_row += 1
            krow = -1
            for kpi in report['kpi'].values():
                if not kpi['enabled']:
                    continue
                krow += 1
                cell = sheet.cell(row=cur_row + krow + 1, column=11)
                cell.value = kpi['fullname']
                color = self.get_color(kpi['name'])
                if color.startswith('#'):
                    color = color[1:]
                cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            cur_row += report['kpi_len']

    def get_calendar(self):
        cal = {}
        sdate = self.request_params.get('date') or datetime.datetime.now().strftime('%m/%Y')
        try:
            cmonth, cyear = sdate.split('/')
        except ValueError:
            cmonth, _, cyear = sdate.split('/')
        cmonth, cyear = int(cmonth), int(cyear)
        firstday = datetime.date(cyear, cmonth, 1)
        nextmonth = firstday + dateutil.relativedelta.relativedelta(months=1)
        if firstday.weekday() == 6:  # Sunday
            curday = firstday
        else:
            curday = firstday - datetime.timedelta(days=self.WDS.index(firstday.weekday()))
        weeks = []
        week = []
        while curday < nextmonth or len(week) > 0:
            day = {'day': curday.day, 'curmonth': curday.month == cmonth, 'week': len(weeks) + 1}
            week.append(day)
            if len(week) == 7:
                weeks.append(week)
                week = []
            curday += datetime.timedelta(days=1)
        cal['firstweek'] = int(firstday.strftime('%W')) + 1
        cal['firstday'] = firstday
        cal['nmonth'] = cmonth
        cal['nyear'] = cyear
        cal['weeks'] = weeks
        cal['weekdays'] = self.WDS
        return cal

    def get_context_data(self):
        context = super().get_context_data()
        if self.is_data_request():
            context['calendar'] = self.get_calendar()
        return context

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()
        sdate = self.request_params.get('date') or datetime.datetime.now().strftime('%m/%Y')
        try:
            sdate = datetime.datetime.strptime(sdate, '%m/%Y')
        except:
            sdate = datetime.datetime.strptime(sdate, '%m/%d/%Y')

        reports = self.filter_reports(chain=True, zone=True, site=True, hours=False, daterange=False)
        reports = reports.filter(datetime__month=sdate.month, datetime__year=sdate.year)
        kpi_sum = self.get_kpis_initial()

        reports = reports.extra(select={'day': "date_part('day', datetime)",
                                        'month': "date_part('month', datetime)",
                                        'year': "date_part('year', datetime)",
                                        })
        agg_reports = self.aggregate_reports(reports, values=['day', 'month', 'year'], order_field='day')

        lastday = sdate.replace(day=1) + dateutil.relativedelta.relativedelta(months=1) - datetime.timedelta(days=1)
        mdays = lastday.day  # last day in month
        data = dict((d, defaultdict(dict)) for d in range(1, mdays + 1))
        curweek = int(sdate.replace(day=1).strftime('%W')) + 1
        day_chartdata = {'name': 'Day-Wise Comparison', 'categories': [], 'series': [], 'xtext': 'WeekDay'}
        day_chartdata['categories'] = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        week_chartdata = {'name': 'Week-Wise Comparison', 'categories': [], 'series': [], 'xtext': 'Week'}
        week_chartdata['categories'] = [curweek]
        month_chartdata = {'name': 'Month At Glance', 'categories': [], 'series': [], 'xtext': 'Days'}
        month_chartdata['categories'] = list(range(1, mdays + 1))
        for row in agg_reports:
            day = int(row['day'])
            wd = self.WDS.index(sdate.replace(day=day).weekday())
            for kname, kpi in list(kpi_sum.items()):
                value = row.get(kpi['id']) or 0
                data[day][kpi['id']] = value
                kpi_sum[kname]['sum'] += value
                if 'weekday' not in kpi_sum[kname]:
                    kpi_sum[kname]['weekday'] = {}
                if wd not in kpi_sum[kname]['weekday']:
                    kpi_sum[kname]['weekday'][wd] = 0
                if 'week' not in kpi_sum[kname]:
                    kpi_sum[kname]['week'] = {}
                if curweek not in kpi_sum[kname]['week']:
                    kpi_sum[kname]['week'][curweek] = 0
                kpi_sum[kname]['weekday'][wd] += value
                kpi_sum[kname]['week'][curweek] += value
            if wd == 6:  # Saturday
                curweek += 1
                week_chartdata['categories'].append(curweek)

        kpi_sum = self.calc_kpis(kpi_sum, reduce=False)
        kpi_sum = self.calc_kpis(kpi_sum, key='week', reduce=False)
        kpi_sum = self.calc_kpis(kpi_sum, key='weekday', reduce=False)

        for kname, kpi in list(kpi_sum.items()):
            if not kpi['enabled']:
                continue
            serie = []
            for w in self.WDS:
                serie.append(kpi_sum[kname]['weekday'].get(w, 0))
            day_chartdata['series'].append({
                'name': self.kpiname(kname), 'type': 'spline', 'data': serie, 'color': self.get_color(kname)
            })
            serie = []
            for c in week_chartdata['categories']:
                serie.append(kpi_sum[kname]['week'].get(c, 0))
            week_chartdata['series'].append({
                'name': self.kpiname(kname), 'type': 'spline', 'data': serie, 'color': self.get_color(kname)
            })
            serie = []
            for d in month_chartdata['categories']:
                serie.append(data.get(d, {}).get(kpi['id'], 0))
            month_chartdata['series'].append({
                'name': self.kpiname(kname), 'type': 'spline', 'data': serie, 'color': self.get_color(kname)
            })

        retval['chartdata_raw'] = [day_chartdata, week_chartdata, month_chartdata]
        retval['chart1data'] = json.dumps(day_chartdata)
        retval['chart2data'] = json.dumps(week_chartdata)
        retval['chart3data'] = json.dumps(month_chartdata)
        retval['data'] = data
        retval['kpi'] = kpi_sum
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])

        return retval


class PerformanceTrendAnalysisReportView(BaseReportView):
    template_name = 'reports/performancetrendanalysis.html'
    data_template_name = 'reports/performancetrendanalysis_data.html'
    pdf_template_name = 'reports/pdf/performancetrendanalysis.html'
    report_name = "Performance Trend Analysis Report"
    export_formats = ['pdf', 'xls', 'csv']
    filter_items = ['Chains', 'Sites', 'Zones', 'Date', 'Hours', 'KPI']

    def write_xls(self, book, context):
        csvdata = context['report']['csvdata']
        sheet = book.active
        cur_row = 1
        default_color = ""
        styles = self.get_xls_styles()
        for cri, csvrow in enumerate(csvdata):
            for cci, csvcell in enumerate(csvrow):
                cell = sheet.cell(row=cur_row + cri, column=cci + 1)
                cell.font = styles['header_font']
                cell.fill = styles['header_fill']
                if cri > 0:
                    kpiname = csvrow[0]
                    color = self.get_color(kpiname, use_reversed=True)
                    if color.startswith('#'):
                        color = color[1:]
                    cell.font = styles['cell_font']
                    cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.value = csvcell

    def get_report_data(self):
        params = self.request_params
        retval = super().get_report_data()

        reports = self.filter_reports(chain=True, zone=True, site=True, hours=True, daterange=False)
        kpi_sum = self.get_kpis_initial()

        current_day = timezone.make_aware(datetime.datetime.strptime(params.get('date'), "%m/%d/%Y"))
        past_day = current_day - datetime.timedelta(days=1)
        past_week = current_day - datetime.timedelta(days=7)
        past_month = current_day - dateutil.relativedelta.relativedelta(months=1)
        past_year = current_day - dateutil.relativedelta.relativedelta(years=1)

        retval['csvdata'].append(['KPIs', 'Current', 'Pastday', '% Pastday',
                                  'Pastweek', '% Pastweek', 'Pastmonth',
                                  '% Pastmonth', 'Pastyear', '% Pastyear'])

        for keyname, rdate in [
                ('current_day', current_day),
                ('past_day', past_day),
                ('past_week', past_week),
                ('past_month', past_month),
                ('past_year', past_year),
        ]:
            drep = reports.filter(
                datetime__range=(timezone.make_aware(datetime.datetime.combine(rdate, datetime.time.min)),
                                 timezone.make_aware(datetime.datetime.combine(rdate, datetime.time.max))))
            drep = drep\
                .extra(select={'wday': "datetime::date"})
            agg_reports = self.aggregate_reports(drep, values=['wday'])
            for kname, kpi in list(kpi_sum.items()):
                kpi_sum[kname][keyname] = 0
                kpi_sum[kname]['%s_perc' % keyname] = "%.02f" % 0
                for row in agg_reports:
                    value = row.get(kpi['id']) or 0
                    kpi_sum[kname][keyname] += value
                    if keyname != 'current_day':
                        variance = 0
                        if value != 0:
                            variance = (kpi_sum[kname]['current_day'] - value) / value
                            variance *= 100
                        kpi_sum[kname]['%s_perc' % keyname] = "%.02f" % variance
                    else:
                        kpi_sum[kname]['sum'] += value

        for key in ('current_day', 'past_day', 'past_week', 'past_month', 'past_year'):
            kpi_sum = self.calc_kpis(kpi_sum, key=key, reduce=False)
        kpi_sum = self.calc_kpis(kpi_sum, reduce=False)

        chartseriesdata = defaultdict(list)
        for kname, kpi in list(kpi_sum.items()):
            if not kpi['enabled']:
                continue
            csvrow = [kpi['fullname']]
            for key in ('current_day', 'past_day', 'past_day_perc',
                        'past_week', 'past_week_perc', 'past_month', 'past_month_perc',
                        'past_year', 'past_year_perc'):
                chartseriesdata[key].append(kpi[key])
                kval = formatkpi(float(kpi[key]), kpi['format'])
                if key.endswith('_perc') and not kval.endswith('%'):
                    kval += "%"
                csvrow.append(kval)
            retval['csvdata'].append(csvrow)

        chartdata = {'categories': [kpi['fullname'] for k, kpi in kpi_sum.items() if kpi['enabled']], 'series': []}
        chartdata['series'] = [
            {'name': 'Current', 'color': '#000000', 'type': 'spline',
                'data': chartseriesdata['current_day']},
            {'name': 'Pastday', 'color': '#00FF00', 'type': 'spline',
                'data': chartseriesdata['past_day']},
            {'name': '% (PD)', 'color': '#ADD9FE', 'type': 'spline',
                'data': chartseriesdata['past_day_perc']},
            {'name': 'PastWeek', 'color': '#FF0000', 'type': 'spline',
                'data': chartseriesdata['past_week']},
            {'name': '% (PW)', 'color': '#FFCCFF', 'type': 'spline',
                'data': chartseriesdata['past_week_perc']},
            {'name': 'PastMonth', 'color': '#FFA500', 'type': 'spline',
                'data': chartseriesdata['past_month']},
            {'name': '% (PM)', 'color': '#529A86', 'type': 'spline',
                'data': chartseriesdata['past_month_perc']},
            {'name': 'Pastyear', 'color': '#A52A2A', 'type': 'spline',
                'data': chartseriesdata['past_year']},
            {'name': '% (PY)', 'color': '#B39EB5', 'type': 'spline',
                'data': chartseriesdata['past_year_perc']},
        ]
        retval['kpi'] = kpi_sum
        retval['chartdata_raw'] = chartdata
        retval['chartdata'] = json.dumps(chartdata)
        retval['kpi_len'] = sum(1 for k in kpi_sum if kpi_sum[k]['enabled'])

        return retval
